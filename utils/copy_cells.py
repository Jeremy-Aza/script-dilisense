import argparse
from openpyxl import load_workbook

def copiar_celdas(range, name_new_sheet):
    file = 'C:/Users/User/Desktop/data_verified_players_2025.xlsx'
    sheet = 'datos'

    wb = load_workbook(file)
    ws_origen = wb[sheet] # o wb['nombre_hoja'] o wb.active

    # Crear nueva hoja
    ws_nueva = wb.create_sheet(name_new_sheet)

    # Definir el rango a copiar
    rango = ws_origen[range]

    # Copiar los valores a la nueva hoja
    for i, fila in enumerate(rango, start=2):
        for j, celda in enumerate(fila, start=1):
            ws_nueva.cell(row=i, column=j, value=celda.value)

    # Guardar el archivo
    wb.save(file)
    print(f"Rango {range} copiado a nueva {name_new_sheet} en el archivo {file}")
    # wb.save('C:/Users/User/Desktop/modified.xlsx')

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Copiar un rango de cedas a una nueva hoja en el mismo archivo excel")
    p.add_argument('--range', type=str, required=True, help='Rango de celdas a copiar')
    p.add_argument('--name_new_sheet', type=str, required=True, help='Nombre de la nueva hoja')
    args = p.parse_args()

    copiar_celdas(range=args.range, name_new_sheet=args.name_new_sheet)

""" copiar_celdas(
    file='C:/Users/User/Desktop/data_verified_players_2025.xlsx',
    sheet_origen='datos', 
    range='A100002:B140001', 
    name_new_sheet='100000a140000'
) """